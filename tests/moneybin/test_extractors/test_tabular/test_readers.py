"""Tests for Stage 2 file readers."""

from pathlib import Path

import polars as pl
import pytest

from moneybin.extractors.tabular.format_detector import FormatInfo
from moneybin.extractors.tabular.readers import read_file


def _write_csv(path: Path, content: str) -> Path:
    path.write_text(content)
    return path


class TestCSVReader:
    """Tests for CSV/text file reading."""

    def test_basic_csv(self, tmp_path: Path) -> None:
        f = _write_csv(
            tmp_path / "basic.csv",
            "Date,Amount,Description\n2026-01-01,42.50,Coffee\n",
        )
        info = FormatInfo(file_type="csv", delimiter=",", encoding="utf-8")
        result = read_file(f, info)
        assert len(result.df) == 1
        assert list(result.df.columns) == ["Date", "Amount", "Description"]

    def test_skip_preamble_rows(self, tmp_path: Path) -> None:
        f = _write_csv(
            tmp_path / "preamble.csv",
            "Bank Summary Report\nGenerated: 2026-01-15\n\nDate,Amount,Description\n2026-01-01,42.50,Coffee\n",
        )
        info = FormatInfo(file_type="csv", delimiter=",", encoding="utf-8")
        result = read_file(f, info)
        assert len(result.df) == 1
        assert "Date" in result.df.columns
        assert result.skip_rows > 0

    def test_trailing_total_row_removed(self, tmp_path: Path) -> None:
        f = _write_csv(
            tmp_path / "trailing.csv",
            "Date,Amount,Description\n2026-01-01,42.50,Coffee\n2026-01-02,10.00,Tea\nTotal,,52.50\n",
        )
        info = FormatInfo(file_type="csv", delimiter=",", encoding="utf-8")
        result = read_file(f, info)
        assert len(result.df) == 2
        assert result.rows_skipped_trailing >= 1

    def test_headerless_csv_keeps_first_data_row(self, tmp_path: Path) -> None:
        """Headerless bank CSV (e.g. Wells Fargo) must not lose row 0.

        WF exports are headerless: ``Date,Amount,*,,Description`` with no
        header line. A real transaction row leads with a date and carries a
        description, so it has a low numeric ratio and the old header
        heuristic mistook row 0 for a header — silently dropping the most
        recent transaction. Regression: all three rows must survive.
        """
        f = _write_csv(
            tmp_path / "wf.csv",
            '"04/16/2026","150.00","*","","RECURRING TRANSFER FROM ACME"\n'
            '"04/15/2026","-150.00","*","","RECURRING TRANSFER TO ACME"\n'
            '"04/14/2026","-50.00","*","","BILL PAY ACME 8230"\n',
        )
        info = FormatInfo(file_type="csv", delimiter=",", encoding="utf-8")
        result = read_file(f, info)
        assert len(result.df) == 3

    def test_summary_row_above_header_not_headerless(self, tmp_path: Path) -> None:
        """A summary/opening-balance line above the real header is preamble.

        A single date+amount row (e.g. an opening-balance line) before the real
        header used to trip headerless detection: the scanner saw row 0 parse
        as a data record and returned ``(0, headerless)`` before reaching the
        ``Date,Amount,Description`` header. Polars then ingested the header as a
        data row under generated column names, breaking downstream mapping.
        The real header must win, with the summary line skipped as preamble.
        """
        f = _write_csv(
            tmp_path / "summary.csv",
            "2026-01-01,100.00\n"
            "Date,Amount,Description\n"
            "2026-01-02,42.50,Coffee\n"
            "2026-01-03,10.00,Tea\n",
        )
        info = FormatInfo(file_type="csv", delimiter=",", encoding="utf-8")
        result = read_file(f, info)
        assert list(result.df.columns) == ["Date", "Amount", "Description"]
        assert len(result.df) == 2
        assert result.skip_rows == 1

    def test_multiple_summary_rows_above_header_not_headerless(
        self, tmp_path: Path
    ) -> None:
        """Several data-like preamble rows above the header are all skipped.

        A statement can carry both an opening-balance and a closing-balance
        line (each a date + amount, so each looks like data) before the real
        header. A single-row peek isn't enough — the scan must look past every
        data-like preamble row and still find the ``Date,Amount,Description``
        header, skipping both summary lines.
        """
        f = _write_csv(
            tmp_path / "two_summary.csv",
            "2026-01-01,100.00\n"
            "2026-01-31,150.00\n"
            "Date,Amount,Description\n"
            "2026-01-02,42.50,Coffee\n"
            "2026-01-03,10.00,Tea\n",
        )
        info = FormatInfo(file_type="csv", delimiter=",", encoding="utf-8")
        result = read_file(f, info)
        assert list(result.df.columns) == ["Date", "Amount", "Description"]
        assert len(result.df) == 2
        assert result.skip_rows == 2

    def test_parenthesized_amount_rows_recognized_in_header_scan(
        self, tmp_path: Path
    ) -> None:
        """Header detection must recognize the amount formats the importer does.

        The loader's ``parse_amount_str`` accepts ``(42.50)`` as a negative, but
        a narrower float check does not. If header detection can't see those as
        amounts, a summary preamble carrying a parenthesized amount above the
        real header is not classified as data, the follow-by-data gate refuses
        the real header, and the file falls back to treating the preamble as the
        header. Recognizing the amount keeps the real header winning.
        """
        f = _write_csv(
            tmp_path / "paren.csv",
            "2026-01-01,(100.00)\n"
            "Date,Amount,Description\n"
            "2026-01-02,(42.50),Refund\n"
            "2026-01-03,(10.00),Fee\n",
        )
        info = FormatInfo(file_type="csv", delimiter=",", encoding="utf-8")
        result = read_file(f, info)
        assert list(result.df.columns) == ["Date", "Amount", "Description"]
        assert len(result.df) == 2
        assert result.skip_rows == 1

    def test_headerless_with_footer_not_mistaken_for_header(
        self, tmp_path: Path
    ) -> None:
        """A date-bearing footer in a headerless file must not become the header.

        A genuinely headerless export (Wells-Fargo-style, every row a
        date+amount) can carry a trailer like ``Downloaded On,2026-04-17``
        within the first 30 lines. That row reads as labels (low numeric ratio)
        and is not a data row (a date but no amount), so a header scan that
        ignores position would pick the footer as the header and skip every
        real data row above it. The file is headerless: the data rows survive
        and the footer is stripped as a trailing row.
        """
        f = _write_csv(
            tmp_path / "wf_footer.csv",
            '"04/16/2026","150.00","*","","RECURRING TRANSFER FROM ACME"\n'
            '"04/15/2026","-150.00","*","","RECURRING TRANSFER TO ACME"\n'
            "Downloaded On,2026-04-17\n",
        )
        info = FormatInfo(file_type="csv", delimiter=",", encoding="utf-8")
        result = read_file(f, info)
        assert result.skip_rows == 0
        assert len(result.df) == 2
        # The most recent transaction (row 0) survived rather than being
        # consumed as a skipped pre-header row.
        assert "04/16/2026" in str(result.df.row(0))

    def test_bom_handled(self, tmp_path: Path) -> None:
        f = tmp_path / "bom.csv"
        f.write_bytes(b"\xef\xbb\xbfDate,Amount\n2026-01-01,42.50\n")
        info = FormatInfo(file_type="csv", delimiter=",", encoding="utf-8-sig")
        result = read_file(f, info)
        assert "Date" in result.df.columns

    def test_pipe_delimiter(self, tmp_path: Path) -> None:
        f = _write_csv(
            tmp_path / "pipe.txt",
            "Date|Amount|Description\n2026-01-01|42.50|Coffee\n",
        )
        info = FormatInfo(file_type="pipe", delimiter="|", encoding="utf-8")
        result = read_file(f, info)
        assert len(result.df) == 1

    def test_row_limit_warning(self, tmp_path: Path) -> None:
        rows = "\n".join(f"2026-01-01,{i},Item{i}" for i in range(10_001))
        f = _write_csv(
            tmp_path / "big.csv",
            f"Date,Amount,Description\n{rows}\n",
        )
        info = FormatInfo(file_type="csv", delimiter=",", encoding="utf-8")
        result = read_file(f, info)
        assert len(result.df) == 10_001
        assert result.row_count_warning is True

    def test_row_limit_refuse(self, tmp_path: Path) -> None:
        rows = "\n".join(f"2026-01-01,{i},Item{i}" for i in range(50_001))
        f = _write_csv(
            tmp_path / "huge.csv",
            f"Date,Amount,Description\n{rows}\n",
        )
        info = FormatInfo(file_type="csv", delimiter=",", encoding="utf-8")
        with pytest.raises(ValueError, match="50,000"):
            read_file(f, info)

    def test_row_limit_refuse_with_override(self, tmp_path: Path) -> None:
        rows = "\n".join(f"2026-01-01,{i},Item{i}" for i in range(50_001))
        f = _write_csv(
            tmp_path / "huge.csv",
            f"Date,Amount,Description\n{rows}\n",
        )
        info = FormatInfo(file_type="csv", delimiter=",", encoding="utf-8")
        result = read_file(f, info, no_row_limit=True)
        assert len(result.df) == 50_001


class TestReadResultTransparency:
    """has_header / rows_in_file / header_row_looks_like_data on ReadResult.

    Regression coverage for the import_preview transparency gap: silent
    header-eating was invisible because ReadResult carried skip_rows and
    rows_skipped_trailing but never surfaced whether a header row was
    consumed, nor a reconcilable total row count.
    """

    def test_has_header_true_for_normal_csv(self, tmp_path: Path) -> None:
        f = _write_csv(
            tmp_path / "basic.csv",
            "Date,Amount,Description\n2026-01-01,42.50,Coffee\n",
        )
        info = FormatInfo(file_type="csv", delimiter=",", encoding="utf-8")
        result = read_file(f, info)
        assert result.has_header is True

    def test_has_header_false_for_headerless_csv(self, tmp_path: Path) -> None:
        f = _write_csv(
            tmp_path / "wf.csv",
            '"04/16/2026","150.00","*","","RECURRING TRANSFER FROM ACME"\n'
            '"04/15/2026","-150.00","*","","RECURRING TRANSFER TO ACME"\n'
            '"04/14/2026","-50.00","*","","BILL PAY ACME 8230"\n',
        )
        info = FormatInfo(file_type="csv", delimiter=",", encoding="utf-8")
        result = read_file(f, info)
        assert result.has_header is False

    def test_rows_in_file_counts_header_plus_data(self, tmp_path: Path) -> None:
        """1-header + 1-data file totals 2 physical rows (property, not stored)."""
        f = _write_csv(
            tmp_path / "basic.csv",
            "Date,Amount,Description\n2026-01-01,42.50,Coffee\n",
        )
        info = FormatInfo(file_type="csv", delimiter=",", encoding="utf-8")
        result = read_file(f, info)
        assert result.rows_in_file == 2  # 1 header + 1 data row
        assert result.has_header is True

    def test_rows_in_file_counts_headerless_data(self, tmp_path: Path) -> None:
        """A 3-row headerless file totals 3 — no phantom header inflates it."""
        f = _write_csv(
            tmp_path / "wf.csv",
            '"04/16/2026","150.00","*","","RECURRING TRANSFER FROM ACME"\n'
            '"04/15/2026","-150.00","*","","RECURRING TRANSFER TO ACME"\n'
            '"04/14/2026","-50.00","*","","BILL PAY ACME 8230"\n',
        )
        info = FormatInfo(file_type="csv", delimiter=",", encoding="utf-8")
        result = read_file(f, info)
        assert result.rows_in_file == 3
        assert result.has_header is False

    def test_rows_in_file_counts_preamble_header_data_trailing(
        self, tmp_path: Path
    ) -> None:
        """3 preamble + 1 header + 2 data + 1 trailing = 7 physical rows."""
        f = _write_csv(
            tmp_path / "combo.csv",
            "Bank Summary Report\nGenerated: 2026-01-15\n\n"
            "Date,Amount,Description\n"
            "2026-01-01,42.50,Coffee\n2026-01-02,10.00,Tea\n"
            "Total,,52.50\n",
        )
        info = FormatInfo(file_type="csv", delimiter=",", encoding="utf-8")
        result = read_file(f, info)
        assert result.skip_rows == 3
        assert result.has_header is True
        assert len(result.df) == 2
        assert result.rows_skipped_trailing == 1
        assert result.rows_in_file == 7

    def test_parquet_has_no_header_and_reconciles(self, tmp_path: Path) -> None:
        """Columnar formats consume no header row: has_header=False, total=len."""
        f = tmp_path / "data.parquet"
        pl.DataFrame({
            "date": ["2026-01-01", "2026-01-02"],
            "amount": ["1.00", "2.00"],
        }).write_parquet(f)
        info = FormatInfo(file_type="parquet", delimiter=None, encoding="utf-8")
        result = read_file(f, info)
        assert result.has_header is False
        assert result.rows_in_file == len(result.df) == 2

    def test_feather_has_no_header_and_reconciles(self, tmp_path: Path) -> None:
        f = tmp_path / "data.feather"
        pl.DataFrame({
            "date": ["2026-01-01", "2026-01-02"],
            "amount": ["1.00", "2.00"],
        }).write_ipc(f)
        info = FormatInfo(file_type="feather", delimiter=None, encoding="utf-8")
        result = read_file(f, info)
        assert result.has_header is False
        assert result.rows_in_file == len(result.df) == 2

    def test_bom_headerless_autodetect_keeps_row0(self, tmp_path: Path) -> None:
        """A BOM'd headerless CSV must auto-detect as headerless on the default path.

        Regression for the review finding: Excel's "CSV UTF-8" export prepends a
        BOM. Opening as utf-8 leaves U+FEFF on physical line 0's first cell,
        which defeated ``detect_date_format`` in ``_detect_header`` — so a
        headerless BOM'd file (row 0 a real transaction) was misread as having a
        header, silently eating the first transaction with NO red flag (the flag
        isn't computed on the auto-detect path). ``_detect_header`` must strip
        the BOM so it correctly returns headerless and keeps every row.
        """
        f = tmp_path / "bom_headerless.csv"
        f.write_bytes(b"\xef\xbb\xbf2026-01-01,42.50,Coffee\n2026-01-02,10.00,Tea\n")
        info = FormatInfo(file_type="csv", delimiter=",", encoding="utf-8-sig")
        result = read_file(f, info)  # no explicit skip_rows — the default path
        assert result.has_header is False
        assert len(result.df) == 2  # neither transaction eaten as a header
        assert result.rows_in_file == 2

    def test_bom_headerless_explicit_skip_still_flags_row0(
        self, tmp_path: Path
    ) -> None:
        """A BOM must not suppress the row-0-is-data red flag on explicit skip.

        A utf-8-sig (BOM) headerless CSV imported with an explicit skip_rows=0
        (wrongly declaring a header) must still flag header_row_looks_like_data:
        the BOM on physical line 0 cannot be allowed to defeat date detection.
        """
        f = tmp_path / "bom.csv"
        f.write_bytes(b"\xef\xbb\xbf2026-01-01,42.50,Coffee\n2026-01-02,10.00,Tea\n")
        info = FormatInfo(file_type="csv", delimiter=",", encoding="utf-8-sig")
        result = read_file(f, info, skip_rows=0)
        assert result.has_header is True
        assert result.header_row_looks_like_data is True

    def test_header_row_looks_like_data_false_for_normal_header(
        self, tmp_path: Path
    ) -> None:
        f = _write_csv(
            tmp_path / "basic.csv",
            "Date,Amount,Description\n2026-01-01,42.50,Coffee\n",
        )
        info = FormatInfo(file_type="csv", delimiter=",", encoding="utf-8")
        result = read_file(f, info)
        assert result.header_row_looks_like_data is False

    def test_header_row_looks_like_data_false_when_headerless(
        self, tmp_path: Path
    ) -> None:
        f = _write_csv(
            tmp_path / "wf.csv",
            '"04/16/2026","150.00","*","","RECURRING TRANSFER FROM ACME"\n'
            '"04/15/2026","-150.00","*","","RECURRING TRANSFER TO ACME"\n',
        )
        info = FormatInfo(file_type="csv", delimiter=",", encoding="utf-8")
        result = read_file(f, info)
        assert result.header_row_looks_like_data is False

    def test_header_row_looks_like_data_true_for_wrong_explicit_skip_rows(
        self, tmp_path: Path
    ) -> None:
        """An explicit skip_rows pointed at a real data row must be flagged.

        Explicit skip_rows always implies has_header=True with no safety
        check today — a caller (human or agent) that mis-specifies skip_rows
        against a genuinely headerless file silently consumes a real
        transaction as a "header". header_row_looks_like_data is the
        defense-in-depth signal for exactly this case.
        """
        f = _write_csv(
            tmp_path / "wf.csv",
            '"04/16/2026","150.00","*","","RECURRING TRANSFER FROM ACME"\n'
            '"04/15/2026","-150.00","*","","RECURRING TRANSFER TO ACME"\n',
        )
        info = FormatInfo(file_type="csv", delimiter=",", encoding="utf-8")
        result = read_file(f, info, skip_rows=0)
        assert result.has_header is True
        assert result.header_row_looks_like_data is True


class TestExcelReader:
    """Tests for Excel file reading."""

    def test_basic_excel(self, tmp_path: Path) -> None:
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["Date", "Amount", "Description"])
        ws.append(["2026-01-01", 42.50, "Coffee"])
        path = tmp_path / "test.xlsx"
        wb.save(path)

        info = FormatInfo(file_type="excel")
        result = read_file(path, info)
        assert len(result.df) == 1
        assert "Date" in result.df.columns

    def test_multi_sheet_picks_largest(self, tmp_path: Path) -> None:
        import openpyxl

        wb = openpyxl.Workbook()
        ws1 = wb.active
        assert ws1 is not None
        ws1.title = "Summary"
        ws1.append(["Total", 100])
        ws2 = wb.create_sheet("Transactions")
        ws2.append(["Date", "Amount", "Desc"])
        ws2.append(["2026-01-01", 42.50, "Coffee"])
        ws2.append(["2026-01-02", 10.00, "Tea"])
        ws2.append(["2026-01-03", 5.00, "Water"])
        path = tmp_path / "multi.xlsx"
        wb.save(path)

        info = FormatInfo(file_type="excel")
        result = read_file(path, info)
        assert len(result.df) == 3

    def test_normal_header_not_flagged(self, tmp_path: Path) -> None:
        """A real Excel header row (labels) must not raise the red flag."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["Date", "Amount", "Description"])
        ws.append(["2026-01-01", 42.50, "Coffee"])
        path = tmp_path / "normal.xlsx"
        wb.save(path)

        result = read_file(path, FormatInfo(file_type="excel"))
        assert result.has_header is True
        assert result.header_row_looks_like_data is False

    def test_headerless_excel_flags_eaten_row0(self, tmp_path: Path) -> None:
        """A headerless Excel sheet's eaten row-0 transaction must be flagged.

        pl.read_excel always consumes row 0 as the header with no headerless
        detection, so a headerless sheet silently loses its first transaction —
        header_row_looks_like_data is the only signal that surfaces it for .xlsx.
        """
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        # No header row — every row is a real transaction (date + amount).
        ws.append(["2026-01-01", 42.50, "Coffee"])
        ws.append(["2026-01-02", 10.00, "Tea"])
        path = tmp_path / "headerless.xlsx"
        wb.save(path)

        result = read_file(path, FormatInfo(file_type="excel"))
        assert result.header_row_looks_like_data is True


class TestParquetReader:
    """Tests for Parquet file reading."""

    def test_basic_parquet(self, tmp_path: Path) -> None:
        df = pl.DataFrame({
            "date": ["2026-01-01"],
            "amount": [42.50],
            "description": ["Coffee"],
        })
        path = tmp_path / "test.parquet"
        df.write_parquet(path)

        info = FormatInfo(file_type="parquet")
        result = read_file(path, info)
        assert len(result.df) == 1
        assert list(result.df.columns) == ["date", "amount", "description"]


class TestFeatherReader:
    """Tests for Feather/Arrow IPC file reading."""

    def test_basic_feather(self, tmp_path: Path) -> None:
        df = pl.DataFrame({
            "date": ["2026-01-01"],
            "amount": [42.50],
            "description": ["Coffee"],
        })
        path = tmp_path / "test.feather"
        df.write_ipc(path)

        info = FormatInfo(file_type="feather")
        result = read_file(path, info)
        assert len(result.df) == 1
