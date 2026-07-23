"""Local export renderer contract tests."""

from __future__ import annotations

import csv
import hashlib
import json
from collections.abc import Sequence
from datetime import UTC, date, datetime
from decimal import Decimal
from pathlib import Path

import duckdb
import pytest
from openpyxl import load_workbook

import moneybin.exports.renderers as renderers
from moneybin.exports.renderers import (
    decode_csv_cell,
    normalize_tabular_cell,
    render_csv,
    render_parquet,
    render_xlsx,
)
from moneybin.exports.snapshot import (
    ExportSubject,
    PreparedColumn,
    PreparedExport,
    PreparedTable,
    ReportExportProvenance,
    build_data_dictionary,
    prepared_table_checksum,
)
from moneybin.privacy.taxonomy import DataClass
from moneybin.tables import TableRef


def make_snapshot(*, report: bool = False) -> PreparedExport:
    columns = (
        PreparedColumn("entry_id", "INTEGER", DataClass.AGGREGATE),
        PreparedColumn("amount", "DECIMAL(18,2)", DataClass.TXN_AMOUNT),
        PreparedColumn("entry_date", "DATE", DataClass.TXN_DATE),
        PreparedColumn("recorded_at", "TIMESTAMP", DataClass.TIMESTAMP_OBSERVABILITY),
        PreparedColumn("note", "VARCHAR", DataClass.DESCRIPTION),
    )
    rows = (
        (
            1,
            Decimal("12.30"),
            date(2026, 7, 20),
            datetime(2026, 7, 20, 8, 9, 10),
            "café, quoted",
        ),
        (2, Decimal("-4.50"), date(2026, 7, 21), None, None),
    )
    table = PreparedTable(
        name="activity",
        source=TableRef("reports", "activity"),
        columns=columns,
        rows=rows,
        checksum_sha256=prepared_table_checksum(columns, rows),
    )
    tables = (table,)
    return PreparedExport(
        artifact_version=1,
        profile="personal",
        created_at=datetime(2026, 7, 21, 18, 42, 33, tzinfo=UTC),
        subject=ExportSubject(
            kind="report" if report else "bundle",
            report_id="test:activity" if report else None,
            parameters={"days": 30} if report else None,
        ),
        redaction_mode="redacted",
        tables=tables,
        data_dictionary=build_data_dictionary(tables),
        provenance=(
            ReportExportProvenance(
                report_id="test:activity",
                receipt={"lineage": ["reports.activity"], "sql": None},
            )
            if report
            else None
        ),
    )


def make_text_snapshot(
    values: Sequence[str | None],
    *,
    column_name: str = "note",
    table_names: Sequence[str] = ("activity",),
) -> PreparedExport:
    columns = (PreparedColumn(column_name, "VARCHAR", DataClass.DESCRIPTION),)
    rows = tuple((value,) for value in values)
    tables = tuple(
        PreparedTable(
            name=name,
            source=TableRef("reports", f"table_{index}"),
            columns=columns,
            rows=rows,
            checksum_sha256=prepared_table_checksum(columns, rows),
        )
        for index, name in enumerate(table_names)
    )
    return PreparedExport(
        artifact_version=1,
        profile="personal",
        created_at=datetime(2026, 7, 21, 18, 42, 33, tzinfo=UTC),
        subject=ExportSubject(kind="bundle"),
        redaction_mode="redacted",
        tables=tables,
        data_dictionary=build_data_dictionary(tables),
        provenance=None,
    )


def test_csv_round_trips_typed_values_through_duckdb(tmp_path: Path) -> None:
    rendered = render_csv(make_snapshot(), tmp_path / "bundle")

    relation = duckdb.read_csv(
        str(rendered.table_files["activity"]),
        header=True,
        columns={
            "entry_id": "INTEGER",
            "amount": "DECIMAL(18,2)",
            "entry_date": "DATE",
            "recorded_at": "TIMESTAMP",
            "note": "VARCHAR",
        },
        na_values="\\N",
    )

    assert relation.columns == [
        "entry_id",
        "amount",
        "entry_date",
        "recorded_at",
        "note",
    ]
    assert relation.fetchall() == list(make_snapshot().tables[0].rows)


def test_normalize_tabular_cell_rejects_unsupported_types() -> None:
    """Unknown prepared values must not silently stringify into an export."""
    with pytest.raises(TypeError, match="Unsupported export cell value"):
        normalize_tabular_cell(object())


def test_csv_null_escape_and_formula_contract_round_trips_losslessly(
    tmp_path: Path,
) -> None:
    values = (
        None,
        r"\N",
        r"\escape",
        "=SUM(A1:A2)",
        "+cmd",
        "-cmd",
        "@cmd",
        "\tcmd",
        "\rcmd",
        "\ncmd",
        "＝cmd",
        "＋cmd",
        "－cmd",
        "＠cmd",
    )
    rendered = render_csv(make_text_snapshot(values), tmp_path / "bundle")

    relation = duckdb.read_csv(
        str(rendered.table_files["activity"]),
        header=True,
        columns={"note": "VARCHAR"},
        na_values=r"\N",
    )
    encoded = [row[0] for row in relation.fetchall()]
    decoded = [decode_csv_cell(value) for value in encoded]
    manifest = json.loads((rendered.path / "manifest.json").read_text())

    assert encoded == [
        None,
        r"\\N",
        r"\\escape",
        r"\=SUM(A1:A2)",
        r"\+cmd",
        r"\-cmd",
        r"\@cmd",
        "\\\tcmd",
        "\\\rcmd",
        "\\\ncmd",
        "\\＝cmd",
        "\\＋cmd",
        "\\－cmd",
        "\\＠cmd",
    ]
    assert decoded == list(values)
    assert manifest["csv_encoding"] == {
        "scheme": "moneybin.csv-cell",
        "version": 2,
        "null": r"\N",
        "escape": "\\",
        "escaped_prefixes": [
            "\\",
            "=",
            "+",
            "-",
            "@",
            "\t",
            "\r",
            "\n",
            "＝",
            "＋",
            "－",
            "＠",
        ],
    }


def test_csv_formula_leading_header_uses_the_reversible_escape_contract(
    tmp_path: Path,
) -> None:
    rendered = render_csv(
        make_text_snapshot(("value",), column_name="=formula"),
        tmp_path / "bundle",
    )

    relation = duckdb.read_csv(
        str(rendered.table_files["activity"]),
        header=True,
        all_varchar=True,
    )

    assert relation.columns == [r"\=formula"]
    assert decode_csv_cell(relation.columns[0]) == "=formula"


def test_csv_added_formula_prefixes_are_reversibly_escaped_in_headers(
    tmp_path: Path,
) -> None:
    prefixes = ("\t", "\r", "\n", "＝", "＋", "－", "＠")

    for index, prefix in enumerate(prefixes):
        rendered = render_csv(
            make_text_snapshot(("value",), column_name=f"{prefix}formula"),
            tmp_path / f"bundle-{index}",
        )
        with rendered.table_files["activity"].open(
            encoding="utf-8", newline=""
        ) as handle:
            encoded_header = next(csv.reader(handle))[0]

        assert encoded_header == f"\\{prefix}formula"
        assert decode_csv_cell(encoded_header) == f"{prefix}formula"


def test_parquet_round_trips_native_typed_values_through_duckdb(
    tmp_path: Path,
) -> None:
    rendered = render_parquet(make_snapshot(), tmp_path / "bundle")

    relation = duckdb.read_parquet(str(rendered.table_files["activity"]))

    assert relation.columns == [
        "entry_id",
        "amount",
        "entry_date",
        "recorded_at",
        "note",
    ]
    assert relation.fetchall() == list(make_snapshot().tables[0].rows)


def test_parquet_rendering_never_uses_duckdb_default_connection(tmp_path: Path) -> None:
    """Rendering an in-memory snapshot must not create a DuckDB spill path."""
    assert not hasattr(renderers, "duckdb")

    render_parquet(make_snapshot(), tmp_path / "bundle")


def test_bundle_manifest_and_sidecars_come_from_the_prepared_snapshot(
    tmp_path: Path,
) -> None:
    snapshot = make_snapshot(report=True)
    rendered = render_csv(snapshot, tmp_path / "bundle")

    manifest = json.loads((rendered.path / "manifest.json").read_text())
    dictionary = json.loads((rendered.path / "data-dictionary.json").read_text())
    table_path = rendered.table_files["activity"]
    table_digest = hashlib.sha256(table_path.read_bytes()).hexdigest()

    assert manifest["created_at"] == "2026-07-21T18:42:33+00:00"
    assert manifest["format"] == "csv"
    assert manifest["redaction_mode"] == "redacted"
    assert manifest["subject"]["report_id"] == "test:activity"
    assert manifest["provenance"]["receipt"]["lineage"] == ["reports.activity"]
    assert manifest["tables"] == [
        {
            "name": "activity",
            "source": "reports.activity",
            "row_count": 2,
            "checksum_sha256": snapshot.tables[0].checksum_sha256,
            "file": "tables/activity.csv",
            "file_checksum_sha256": table_digest,
            "columns": [
                {
                    "name": column.name,
                    "duckdb_type": column.duckdb_type,
                    "data_class": column.data_class.value,
                }
                for column in snapshot.tables[0].columns
            ],
        }
    ]
    assert dictionary == snapshot.data_dictionary
    assert (rendered.path / "checksums.sha256").read_text() == (
        f"{table_digest}  tables/activity.csv\n"
    )


def test_xlsx_contains_data_and_visible_receipt_sheets(tmp_path: Path) -> None:
    rendered = render_xlsx(make_snapshot(report=True), tmp_path)

    workbook = load_workbook(rendered.path, data_only=True)

    assert workbook.sheetnames == [
        "activity",
        "MoneyBin Manifest",
        "MoneyBin Data Dictionary",
    ]
    assert all(workbook[name].sheet_state == "visible" for name in workbook.sheetnames)
    assert list(workbook["activity"].values) == [
        ("entry_id", "amount", "entry_date", "recorded_at", "note"),
        (1, "12.30", "2026-07-20", "2026-07-20T08:09:10", "café, quoted"),
        (2, "-4.50", "2026-07-21", None, None),
    ]
    manifest = json.loads(workbook["MoneyBin Manifest"]["A2"].value)
    dictionary = json.loads(workbook["MoneyBin Data Dictionary"]["A2"].value)
    assert manifest["format"] == "xlsx"
    assert manifest["tables"][0]["worksheet"] == "activity"
    assert manifest["tables"][0]["checksum_sha256"] == (
        make_snapshot().tables[0].checksum_sha256
    )
    assert dictionary == make_snapshot().data_dictionary


def test_xlsx_formula_leading_strings_remain_exact_literal_text(tmp_path: Path) -> None:
    values = ("=SUM(A1:A2)", "+cmd", "-cmd", "@cmd")
    rendered = render_xlsx(make_text_snapshot(values), tmp_path)

    workbook = load_workbook(rendered.path, data_only=False)
    cells = list(workbook["activity"]["A"])[1:]

    assert [cell.value for cell in cells] == list(values)
    assert [cell.data_type for cell in cells] == ["s", "s", "s", "s"]


def test_xlsx_worksheet_names_avoid_case_insensitive_and_receipt_collisions(
    tmp_path: Path,
) -> None:
    snapshot = make_text_snapshot(
        ("value",),
        table_names=("activity", "ACTIVITY", "MoneyBin Manifest"),
    )

    rendered = render_xlsx(snapshot, tmp_path)
    workbook = load_workbook(rendered.path, data_only=True)

    assert workbook.sheetnames == [
        "activity",
        "ACTIVITY-2",
        "MoneyBin Manifest-2",
        "MoneyBin Manifest",
        "MoneyBin Data Dictionary",
    ]
    assert len({name.casefold() for name in workbook.sheetnames}) == len(
        workbook.sheetnames
    )
