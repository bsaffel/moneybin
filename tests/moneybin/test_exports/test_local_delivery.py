"""Atomic local export publication tests."""

from __future__ import annotations

import hashlib
import json
import os
import stat
from dataclasses import replace
from pathlib import Path
from zipfile import ZipFile

import pytest
from openpyxl import load_workbook

import moneybin.exports.local as local_delivery
from moneybin.exports.local import LocalExportPublisher
from moneybin.exports.manifest import LocalExportFormat
from moneybin.exports.renderers import RenderedArtifact
from moneybin.exports.snapshot import PreparedExport
from moneybin.services.request_lifetime import (
    PublicationCancelledError,
    RequestLifetime,
)
from tests.moneybin.test_exports.test_renderers import make_snapshot, make_text_snapshot


def _mode(path: Path) -> int:
    return stat.S_IMODE(path.stat().st_mode)


def _digest(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


@pytest.mark.parametrize(
    ("format", "extension"), [("csv", "csv"), ("parquet", "parquet")]
)
def test_publish_creates_an_exact_restrictive_timestamped_bundle(
    tmp_path: Path,
    format: LocalExportFormat,
    extension: str,
) -> None:
    exports_root = tmp_path / "exports"
    publisher = LocalExportPublisher(exports_root)

    receipt = publisher.publish(make_snapshot(), format=format, compress_zip=False)

    assert receipt.artifact_path is not None
    assert receipt.artifact_path.is_absolute()
    assert receipt.artifact_path.name == "export-20260721T184233Z"
    assert receipt.destination.local_path == exports_root.resolve()
    assert receipt.subject == {"kind": "bundle"}
    assert receipt.redaction_mode == "redacted"
    assert receipt.row_counts == {"activity": 2}
    assert receipt.compressed_artifact_path is None
    assert receipt.recovery_actions == ()
    assert {
        path.relative_to(receipt.artifact_path).as_posix()
        for path in receipt.artifact_path.rglob("*")
    } == {
        "manifest.json",
        "checksums.sha256",
        "data-dictionary.json",
        "tables",
        f"tables/activity.{extension}",
    }
    assert _mode(receipt.artifact_path) == 0o700
    assert _mode(receipt.artifact_path / "tables") == 0o700
    assert all(
        _mode(path) == 0o600
        for path in receipt.artifact_path.rglob("*")
        if path.is_file()
    )
    assert receipt.checksums == {
        path.relative_to(receipt.artifact_path).as_posix(): _digest(path)
        for path in receipt.artifact_path.rglob("*")
        if path.is_file()
    }


@pytest.mark.parametrize("format", ["csv", "parquet"])
def test_empty_snapshot_publishes_required_restrictive_bundle_layout(
    tmp_path: Path,
    format: LocalExportFormat,
) -> None:
    receipt = LocalExportPublisher(tmp_path / "exports").publish(
        make_text_snapshot((), table_names=()),
        format=format,
        compress_zip=False,
    )

    assert receipt.artifact_path is not None
    assert receipt.row_counts == {}
    assert {
        path.relative_to(receipt.artifact_path).as_posix()
        for path in receipt.artifact_path.rglob("*")
    } == {
        "manifest.json",
        "checksums.sha256",
        "data-dictionary.json",
        "tables",
    }
    assert _mode(receipt.artifact_path) == 0o700
    assert _mode(receipt.artifact_path / "tables") == 0o700
    assert all(
        _mode(receipt.artifact_path / name) == 0o600
        for name in ("manifest.json", "checksums.sha256", "data-dictionary.json")
    )
    assert receipt.checksums == {
        name: _digest(receipt.artifact_path / name)
        for name in ("manifest.json", "checksums.sha256", "data-dictionary.json")
    }
    manifest = json.loads((receipt.artifact_path / "manifest.json").read_text())
    assert manifest["tables"] == []
    assert (receipt.artifact_path / "checksums.sha256").read_text() == ""


def test_publish_never_overwrites_a_successful_collision(tmp_path: Path) -> None:
    exports_root = tmp_path / "exports"
    publisher = LocalExportPublisher(exports_root)

    first = publisher.publish(make_snapshot(), format="csv", compress_zip=False)
    second = publisher.publish(make_snapshot(), format="csv", compress_zip=False)

    assert first.artifact_path is not None
    assert second.artifact_path is not None
    assert first.artifact_path.name == "export-20260721T184233Z"
    assert second.artifact_path.name == "export-20260721T184233Z-2"
    assert first.artifact_path.exists()
    assert second.artifact_path.exists()
    assert _mode(exports_root / ".publish.lock") == 0o600


def test_cancelled_local_publish_cleans_staging_before_visible_output(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Cancellation after rendering cannot leave a partial visible bundle."""
    exports_root = tmp_path / "exports"
    publisher = LocalExportPublisher(exports_root)
    lifetime = RequestLifetime()
    original_render = publisher._render  # pyright: ignore[reportPrivateUsage]

    def render_then_cancel(
        snapshot: PreparedExport,
        format: LocalExportFormat,
        staging_root: Path,
    ) -> Path:
        rendered = original_render(snapshot, format, staging_root)
        lifetime.cancel_and_wait()
        return rendered

    monkeypatch.setattr(publisher, "_render", render_then_cancel)

    with pytest.raises(PublicationCancelledError):
        publisher.publish(
            make_snapshot(),
            format="csv",
            compress_zip=False,
            publication_lifetime=lifetime,
        )

    assert not list(exports_root.glob("export-*"))
    assert not list(exports_root.glob(".staging-*"))


def test_local_destination_file_collision_is_rejected(tmp_path: Path) -> None:
    """A configured file cannot be mistaken for an export directory."""
    exports_root = tmp_path / "not-a-directory"
    exports_root.write_text("occupied")

    with pytest.raises(ValueError, match="not a directory"):
        LocalExportPublisher(exports_root).publish(
            make_snapshot(), format="csv", compress_zip=False
        )


def test_tampered_staging_bundle_is_not_published_or_left_behind(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    exports_root = tmp_path / "exports"
    original = local_delivery.render_csv

    def render_then_tamper(
        snapshot: PreparedExport, staging_root: Path
    ) -> RenderedArtifact:
        rendered = original(snapshot, staging_root)
        rendered.table_files["activity"].write_bytes(b"tampered")
        return rendered

    monkeypatch.setattr(local_delivery, "render_csv", render_then_tamper)

    with pytest.raises(ValueError, match="checksum"):
        LocalExportPublisher(exports_root).publish(
            make_snapshot(), format="csv", compress_zip=False
        )

    assert not list(exports_root.glob("export-*"))
    assert not list(exports_root.glob(".staging-*"))


@pytest.mark.parametrize("format", ["csv", "parquet"])
@pytest.mark.parametrize(
    ("drift", "failure"),
    [("rows", "cell"), ("columns", "column")],
)
def test_renderer_content_drift_is_not_published(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    format: LocalExportFormat,
    drift: str,
    failure: str,
) -> None:
    """Receipts cannot make a renderer payload drift from the snapshot publishable."""
    snapshot = make_snapshot()
    table = snapshot.tables[0]
    wrong_table = (
        replace(table, rows=((999, *table.rows[0][1:]), *table.rows[1:]))
        if drift == "rows"
        else replace(
            table,
            columns=(
                replace(table.columns[0], name="wrong_entry_id"),
                *table.columns[1:],
            ),
        )
    )
    rendered_snapshot = replace(
        snapshot,
        tables=(wrong_table,),
    )
    renderer = (
        local_delivery.render_csv if format == "csv" else local_delivery.render_parquet
    )

    def render_wrong_snapshot(
        _snapshot: PreparedExport, staging_root: Path
    ) -> RenderedArtifact:
        return renderer(rendered_snapshot, staging_root)

    monkeypatch.setattr(local_delivery, f"render_{format}", render_wrong_snapshot)

    with pytest.raises(ValueError, match=failure):
        LocalExportPublisher(tmp_path / "exports").publish(
            snapshot, format=format, compress_zip=False
        )

    assert not list((tmp_path / "exports").glob("export-*"))
    assert not list((tmp_path / "exports").glob(".staging-*"))


def test_unexpected_bundle_directory_is_not_published(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    exports_root = tmp_path / "exports"
    original = local_delivery.render_csv

    def render_with_directory(
        snapshot: PreparedExport, staging_root: Path
    ) -> RenderedArtifact:
        rendered = original(snapshot, staging_root)
        (rendered.path / "unexpected").mkdir()
        return rendered

    monkeypatch.setattr(local_delivery, "render_csv", render_with_directory)

    with pytest.raises(ValueError, match="layout"):
        LocalExportPublisher(exports_root).publish(
            make_snapshot(), format="csv", compress_zip=False
        )

    assert not list(exports_root.glob("export-*"))
    assert not list(exports_root.glob(".staging-*"))


def test_symlinked_bundle_directory_is_rejected_before_chmod(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    exports_root = tmp_path / "exports"
    publisher = LocalExportPublisher(exports_root)
    existing = publisher.publish(make_snapshot(), format="csv", compress_zip=False)
    assert existing.artifact_path is not None
    existing_manifest = (existing.artifact_path / "manifest.json").read_bytes()

    outside = tmp_path / "outside"
    outside.mkdir(mode=0o751)
    outside.chmod(0o751)
    original = local_delivery.render_csv

    def render_with_symlink(
        snapshot: PreparedExport, staging_root: Path
    ) -> RenderedArtifact:
        rendered = original(snapshot, staging_root)
        (rendered.path / "unexpected").symlink_to(outside, target_is_directory=True)
        return rendered

    monkeypatch.setattr(local_delivery, "render_csv", render_with_symlink)

    with pytest.raises(ValueError, match="symlink"):
        publisher.publish(make_snapshot(), format="csv", compress_zip=False)

    assert _mode(outside) == 0o751
    assert existing.artifact_path.exists()
    assert (existing.artifact_path / "manifest.json").read_bytes() == existing_manifest
    assert {path.name for path in exports_root.glob("export-*")} == {
        "export-20260721T184233Z"
    }
    assert not list(exports_root.glob(".staging-*"))


def test_symlinked_bundle_file_is_rejected(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    exports_root = tmp_path / "exports"
    outside = tmp_path / "outside.csv"
    original = local_delivery.render_csv

    def render_with_symlink(
        snapshot: PreparedExport, staging_root: Path
    ) -> RenderedArtifact:
        rendered = original(snapshot, staging_root)
        table_path = rendered.table_files["activity"]
        outside.write_bytes(table_path.read_bytes())
        table_path.unlink()
        table_path.symlink_to(outside)
        return rendered

    monkeypatch.setattr(local_delivery, "render_csv", render_with_symlink)

    with pytest.raises(ValueError, match="symlink"):
        LocalExportPublisher(exports_root).publish(
            make_snapshot(), format="csv", compress_zip=False
        )

    assert outside.exists()
    assert not list(exports_root.glob("export-*"))
    assert not list(exports_root.glob(".staging-*"))


def test_snapshot_manifest_tamper_is_not_published(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    exports_root = tmp_path / "exports"
    original = local_delivery.render_csv

    def render_then_tamper(
        snapshot: PreparedExport, staging_root: Path
    ) -> RenderedArtifact:
        rendered = original(snapshot, staging_root)
        manifest_path = rendered.path / "manifest.json"
        manifest = json.loads(manifest_path.read_text())
        manifest["profile"] = "other"
        manifest["subject"] = {"kind": "report", "report_id": "other"}
        manifest["provenance"] = {"report_id": "other", "receipt": {}}
        manifest_path.write_text(json.dumps(manifest))
        return rendered

    monkeypatch.setattr(local_delivery, "render_csv", render_then_tamper)

    with pytest.raises(ValueError, match="manifest"):
        LocalExportPublisher(exports_root).publish(
            make_snapshot(report=True), format="csv", compress_zip=False
        )

    assert not list(exports_root.glob("export-*"))


def test_internally_consistent_dictionary_tamper_is_not_published(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    exports_root = tmp_path / "exports"
    original = local_delivery.render_csv

    def render_then_tamper(
        snapshot: PreparedExport, staging_root: Path
    ) -> RenderedArtifact:
        rendered = original(snapshot, staging_root)
        manifest_path = rendered.path / "manifest.json"
        dictionary_path = rendered.path / "data-dictionary.json"
        manifest = json.loads(manifest_path.read_text())
        dictionary = json.loads(dictionary_path.read_text())
        dictionary["tables"][0]["source"] = "reports.other"
        manifest["data_dictionary"] = dictionary
        manifest_path.write_text(json.dumps(manifest))
        dictionary_path.write_text(json.dumps(dictionary))
        return rendered

    monkeypatch.setattr(local_delivery, "render_csv", render_then_tamper)

    with pytest.raises(ValueError, match="dictionary"):
        LocalExportPublisher(exports_root).publish(
            make_snapshot(), format="csv", compress_zip=False
        )

    assert not list(exports_root.glob("export-*"))


def test_receipt_uses_independently_verified_bytes_not_renderer_fields(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    original = local_delivery.render_csv

    def render_with_false_receipt(
        snapshot: PreparedExport, staging_root: Path
    ) -> RenderedArtifact:
        rendered = original(snapshot, staging_root)
        return replace(rendered, file_checksums={"unverified": "0" * 64})

    monkeypatch.setattr(local_delivery, "render_csv", render_with_false_receipt)

    receipt = LocalExportPublisher(tmp_path / "exports").publish(
        make_snapshot(), format="csv", compress_zip=False
    )

    assert receipt.artifact_path is not None
    assert "unverified" not in receipt.checksums
    assert receipt.checksums == {
        path.relative_to(receipt.artifact_path).as_posix(): _digest(path)
        for path in receipt.artifact_path.rglob("*")
        if path.is_file()
    }


def test_zip_is_an_additional_complete_validated_bundle(tmp_path: Path) -> None:
    receipt = LocalExportPublisher(tmp_path / "exports").publish(
        make_snapshot(), format="csv", compress_zip=True
    )

    assert receipt.artifact_path is not None
    assert receipt.artifact_path.is_dir()
    assert receipt.compressed_artifact_path is not None
    assert receipt.compressed_artifact_path.name == f"{receipt.artifact_path.name}.zip"
    assert _mode(receipt.compressed_artifact_path) == 0o600
    with ZipFile(receipt.compressed_artifact_path) as archive:
        assert set(archive.namelist()) == {
            "manifest.json",
            "checksums.sha256",
            "data-dictionary.json",
            "tables/activity.csv",
        }
        for member in archive.namelist():
            assert archive.read(member) == (receipt.artifact_path / member).read_bytes()
    assert receipt.checksums["archive.zip"] == _digest(receipt.compressed_artifact_path)


def test_zip_rename_failure_rolls_back_only_this_publication(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    exports_root = tmp_path / "exports"
    publisher = LocalExportPublisher(exports_root)
    existing = publisher.publish(make_snapshot(), format="csv", compress_zip=False)
    original_rename = Path.rename

    def fail_zip_rename(path: Path, target: Path) -> Path:
        if path.name == "artifact.zip":
            raise OSError("injected ZIP rename failure")
        return original_rename(path, target)

    monkeypatch.setattr(Path, "rename", fail_zip_rename)

    with pytest.raises(OSError, match="injected ZIP"):
        publisher.publish(make_snapshot(), format="csv", compress_zip=True)

    assert existing.artifact_path is not None
    assert existing.artifact_path.exists()
    assert {path.name for path in exports_root.glob("export-*")} == {
        "export-20260721T184233Z"
    }
    assert not list(exports_root.glob(".staging-*"))


def test_xlsx_rejects_zip_before_creating_staging(tmp_path: Path) -> None:
    exports_root = tmp_path / "exports"

    with pytest.raises(ValueError, match="XLSX"):
        LocalExportPublisher(exports_root).publish(
            make_snapshot(), format="xlsx", compress_zip=True
        )

    assert not exports_root.exists()


def test_tampered_xlsx_data_is_not_published(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    exports_root = tmp_path / "exports"
    original = local_delivery.render_xlsx

    def render_then_tamper(
        snapshot: PreparedExport, staging_root: Path
    ) -> RenderedArtifact:
        rendered = original(snapshot, staging_root)
        workbook = load_workbook(rendered.path)
        workbook["activity"]["B2"] = "999.99"
        workbook.save(rendered.path)
        return rendered

    monkeypatch.setattr(local_delivery, "render_xlsx", render_then_tamper)

    with pytest.raises(ValueError, match="cell"):
        LocalExportPublisher(exports_root).publish(
            make_snapshot(), format="xlsx", compress_zip=False
        )

    assert not list(exports_root.glob("export-*"))
    assert not list(exports_root.glob(".staging-*"))


def test_tampered_xlsx_metadata_is_not_published(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    exports_root = tmp_path / "exports"
    original = local_delivery.render_xlsx

    def render_then_tamper(
        snapshot: PreparedExport, staging_root: Path
    ) -> RenderedArtifact:
        rendered = original(snapshot, staging_root)
        workbook = load_workbook(rendered.path)
        manifest = json.loads(workbook["MoneyBin Manifest"]["A2"].value)
        manifest["redaction_mode"] = "unredacted"
        workbook["MoneyBin Manifest"]["A2"] = json.dumps(manifest)
        workbook.save(rendered.path)
        return rendered

    monkeypatch.setattr(local_delivery, "render_xlsx", render_then_tamper)

    with pytest.raises(ValueError, match="manifest"):
        LocalExportPublisher(exports_root).publish(
            make_snapshot(report=True), format="xlsx", compress_zip=False
        )

    assert not list(exports_root.glob("export-*"))


def test_xlsx_publishes_one_restrictive_timestamped_workbook(tmp_path: Path) -> None:
    receipt = LocalExportPublisher(tmp_path / "exports").publish(
        make_snapshot(), format="xlsx", compress_zip=False
    )

    assert receipt.artifact_path is not None
    assert receipt.artifact_path.name == "export-20260721T184233Z.xlsx"
    assert receipt.artifact_path.is_file()
    assert _mode(receipt.artifact_path) == 0o600
    assert receipt.checksums == {"export.xlsx": _digest(receipt.artifact_path)}


def test_outer_staging_mode_is_exact_under_restrictive_umask(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    observed_modes: list[int] = []
    original = local_delivery.render_csv

    def inspect_stage_then_render(
        snapshot: PreparedExport, staging_root: Path
    ) -> RenderedArtifact:
        observed_modes.append(_mode(staging_root.parent))
        return original(snapshot, staging_root)

    monkeypatch.setattr(local_delivery, "render_csv", inspect_stage_then_render)
    previous_umask = os.umask(0o777)
    try:
        LocalExportPublisher(tmp_path / "exports").publish(
            make_snapshot(), format="csv", compress_zip=False
        )
    finally:
        os.umask(previous_umask)

    assert observed_modes == [0o700]
