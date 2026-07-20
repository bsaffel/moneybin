"""Stage 2: File readers producing format-agnostic Polars DataFrames.

Each reader converts a specific file type into a Polars DataFrame with
string column names. This is the format-agnostic boundary — everything
downstream operates on DataFrames regardless of source format.
"""

import logging
import re
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path

import polars as pl

from moneybin.extractors.tabular.date_detection import (
    detect_date_format,
    parse_amount_str,
)
from moneybin.extractors.tabular.format_detector import (
    FormatInfo,
    _read_sample_lines,  # pyright: ignore[reportPrivateUsage]  # shared package helper
)

logger = logging.getLogger(__name__)


DEFAULT_TRAILING_PATTERNS: list[str] = [
    r"^(Total|Grand Total|Sum|Totals)\b",
    r"^(Export(ed)?|Generated|Downloaded|Report) (Date|On|At)\b",
    r"^(Record Count|Row Count|Number of)",
    r"^(Opening|Closing|Beginning|Ending) Balance\b",
    r"^,{3,}$",
    r"^\s*$",
]


@dataclass
class ReadResult:
    """Output of a file reader."""

    df: pl.DataFrame
    skip_rows: int = 0
    rows_skipped_trailing: int = 0
    row_count_warning: bool = False
    sheet_used: str | None = None
    has_header: bool = True
    """Whether a header row was consumed. False for a detected-headerless file
    and for schema-typed formats (parquet/feather) where column names are
    metadata, not a consumed row."""
    header_row_looks_like_data: bool = False
    """True when the row consumed as the header also parses as a transaction
    record (date + amount) — a red flag that a real data row may have been
    eaten as a header via an explicit skip_rows (CSV) or the always-on header
    assumption (Excel). CSV auto-detection is its own safety net (it never
    picks a data-looking row as the header), so this stays False there."""

    @property
    def rows_in_file(self) -> int:
        """Reconciled source-row accounting derived from the reader's fields.

        Sums preamble skipped + header (0/1) + rows loaded + trailing rows
        dropped. Equals the file's physical row count in the common case, but
        is derived from the reader's own accounting — NOT an independent
        physical line count. It can undercount the raw file where the parser
        coalesced rows the reader never saw as separate (repeated-header dedup
        in paginated exports, quoted fields containing embedded newlines). Use
        it to reconcile the reader's accounting, not as a data-loss oracle.
        """
        return (
            self.skip_rows
            + (1 if self.has_header else 0)
            + len(self.df)
            + self.rows_skipped_trailing
        )


def read_file(
    path: Path,
    info: FormatInfo,
    *,
    skip_rows: int | None = None,
    sheet: str | None = None,
    skip_trailing_patterns: list[str] | None = None,
    no_row_limit: bool = False,
    source_bytes: bytes | None = None,
    has_header: bool | None = None,
) -> ReadResult:
    """Read a file into a format-agnostic Polars DataFrame.

    Args:
        path: File path.
        info: Format detection result from Stage 1.
        skip_rows: Explicit skip rows (overrides detection).
        sheet: Excel sheet name (overrides auto-selection).
        skip_trailing_patterns: Regex patterns for trailing junk.
            None = use defaults, [] = no patterns.
        no_row_limit: If True, skip row count limits.
        source_bytes: Already materialized source object to parse.
        has_header: Persisted header decision; None runs detection.

    Returns:
        ReadResult with DataFrame and metadata.

    Raises:
        ValueError: If row count exceeds limit without override.
    """
    if info.file_type in ("csv", "tsv", "pipe", "semicolon"):
        result = _read_text(
            path,
            info,
            skip_rows=skip_rows,
            skip_trailing_patterns=skip_trailing_patterns,
            source_bytes=source_bytes,
            has_header=has_header,
        )
    elif info.file_type == "excel":
        result = _read_excel(
            path,
            info,
            skip_rows=skip_rows,
            sheet=sheet,
            source_bytes=source_bytes,
        )
    elif info.file_type == "parquet":
        result = _read_parquet(path, source_bytes=source_bytes)
    elif info.file_type == "feather":
        result = _read_feather(path, source_bytes=source_bytes)
    else:
        raise ValueError(f"No reader for file type: {info.file_type}")

    from moneybin.config import get_settings

    tabular_cfg = get_settings().providers.tabular
    row_count = len(result.df)
    if row_count > tabular_cfg.row_refuse_threshold and not no_row_limit:
        raise ValueError(
            f"File has {row_count:,} rows, exceeding the "
            f"{tabular_cfg.row_refuse_threshold:,} row limit. "
            f"Use --no-row-limit to override."
        )
    if row_count > tabular_cfg.row_warn_threshold:
        logger.warning(
            f"⚠️  File has {row_count:,} rows (warning threshold: "
            f"{tabular_cfg.row_warn_threshold:,}). Proceeding with import."
        )
        result.row_count_warning = True

    return result


def _read_text(
    path: Path,
    info: FormatInfo,
    *,
    skip_rows: int | None = None,
    skip_trailing_patterns: list[str] | None = None,
    source_bytes: bytes | None = None,
    has_header: bool | None = None,
) -> ReadResult:
    """Read a text-based tabular file (CSV, TSV, pipe, semicolon).

    Args:
        path: File path.
        info: Format detection result with delimiter and encoding.
        skip_rows: Explicit preamble rows to skip (overrides auto-detection).
        skip_trailing_patterns: Regex patterns for trailing junk rows.
        source_bytes: Already materialized source object to parse.
        has_header: Persisted header decision; None runs detection.

    Returns:
        ReadResult with the parsed DataFrame and metadata.
    """
    encoding = info.encoding
    delimiter = info.delimiter or ","

    # Explicit skip_rows implies a header at that row; auto-detection both
    # locates the header and decides whether the file has one at all.
    explicit_skip = skip_rows is not None
    resolved_has_header = True
    if skip_rows is None:
        skip_rows, resolved_has_header = _detect_header(
            path,
            encoding,
            delimiter,
            source_bytes=source_bytes,
        )
    elif has_header is not None:
        resolved_has_header = has_header

    # header_row_looks_like_data is defense-in-depth for the EXPLICIT skip_rows
    # path only (has_header is unconditionally True there — no safety check of
    # its own). Auto-detection (_detect_header) never selects a data-looking row
    # as the header, so computing it there would always be False — skip the read.
    header_row_looks_like_data = False
    if explicit_skip:
        header_row_looks_like_data = resolved_has_header and _row_looks_like_data_at(
            path,
            encoding,
            delimiter,
            skip_rows,
            source_bytes=source_bytes,
        )

    df = pl.read_csv(
        path if source_bytes is None else BytesIO(source_bytes),
        separator=delimiter,
        encoding=encoding if encoding != "utf-8-sig" else "utf8",
        skip_rows=skip_rows,
        has_header=resolved_has_header,
        infer_schema_length=0,
        truncate_ragged_lines=True,
    )

    patterns = skip_trailing_patterns
    if patterns is None:
        patterns = DEFAULT_TRAILING_PATTERNS
    rows_removed = 0
    if patterns and len(df) > 0:
        df, rows_removed = _remove_trailing_rows(df, patterns)

    if len(df) > 0:
        df = _remove_repeated_headers(df)

    return ReadResult(
        df=df,
        skip_rows=skip_rows,
        rows_skipped_trailing=rows_removed,
        has_header=resolved_has_header,
        header_row_looks_like_data=header_row_looks_like_data,
    )


def _detect_header(
    path: Path,
    encoding: str,
    delimiter: str,
    *,
    source_bytes: bytes | None = None,
) -> tuple[int, bool]:
    """Locate the header row, or determine the file is headerless.

    Scans up to the first 30 content rows and decides between two outcomes:

    - **Header present.** The first row that reads as labels (low numeric
      ratio), does *not* itself parse as a transaction, *and* is followed by
      a data row is the header. Returns ``(row_index, True)``. Scanning the
      whole window means any number of data-like preamble rows above the
      header — opening- and closing-balance summary lines such as
      ``2026-01-01,100.00`` — are skipped rather than mistaken for the first
      row of a headerless file. The follow-by-data check is what keeps a
      footer/trailer that also reads as labels (``Downloaded On,2026-04-17``,
      sitting *below* the data in a headerless file) from winning.
    - **Headerless.** When no row qualifies as a header, the first row that
      parses as a data record (date plus numeric amount) starts the data.
      Returns ``(row_index, False)`` so the reader keeps that row. This is
      the Wells Fargo case: ``Date,Amount,*,,Description`` with no header
      line, where every row leads with a date (low numeric ratio) and so
      none reads as a header.

    Args:
        path: File path.
        encoding: File encoding.
        delimiter: Column delimiter.
        source_bytes: Already materialized source object to inspect.

    Returns:
        ``(skip_rows, has_header)`` — rows to skip before the header (or
        before the first data row when headerless), and whether a header
        row is present.
    """
    enc = encoding if encoding != "utf-8-sig" else "utf-8"
    lines = [
        line.lstrip("\ufeff")
        for line in _read_sample_lines(
            path,
            enc,
            n=30,
            source_bytes=source_bytes,
        )
    ]

    # Two passes (see docstring): find a label row followed by data, else fall
    # back to the first data row as headerless.
    qualifying: list[tuple[int, list[str]]] = []
    for i, line in enumerate(lines):
        if not line.strip():
            continue
        parts = line.split(delimiter)
        if len(parts) < 2:
            continue
        non_empty = [p.strip().strip('"').strip("'") for p in parts if p.strip()]
        if not non_empty:
            continue
        qualifying.append((i, non_empty))

    for idx, (i, non_empty) in enumerate(qualifying):
        if _looks_like_header(non_empty) and not _looks_like_data_row(non_empty):
            # A real header sits above the data, so a data row must follow it.
            # This rejects a footer/trailer that also reads as labels (e.g.
            # "Downloaded On,2026-04-17": a date, no amount, low numeric ratio)
            # but appears after the data in a headerless file — without the
            # follow check it would win header detection and the rows above it
            # would be skipped as preamble.
            if any(_looks_like_data_row(later) for _, later in qualifying[idx + 1 :]):
                return i, True
    for i, non_empty in qualifying:
        if _looks_like_data_row(non_empty):
            return i, False

    return 0, True


def _row_looks_like_data_at(
    path: Path,
    encoding: str,
    delimiter: str,
    row_index: int,
    *,
    source_bytes: bytes | None = None,
) -> bool:
    """Return True if the physical row at ``row_index`` parses as a transaction.

    Defense-in-depth check on the row an explicit ``skip_rows`` override is
    about to consume as a header — that path sets ``has_header=True`` with no
    safety check of its own (unlike auto-detection). A row that both is treated
    as a header AND parses as a data record is a red flag: a real transaction
    may be about to be silently dropped.

    Args:
        path: File path.
        encoding: File encoding.
        delimiter: Column delimiter.
        row_index: Zero-based physical line index to check.
        source_bytes: Already materialized source object to inspect.

    Returns:
        True when the row at ``row_index`` parses as a transaction record.
    """
    enc = encoding if encoding != "utf-8-sig" else "utf-8"
    lines = _read_sample_lines(
        path,
        enc,
        n=row_index + 1,
        source_bytes=source_bytes,
    )
    if row_index >= len(lines):
        return False
    # lstrip a leading BOM: a utf-8-sig file is decoded as utf-8 here
    # (polars-compatible), so physical line 0 may retain it.
    parts = lines[row_index].lstrip("\ufeff").split(delimiter)
    non_empty = [p.strip().strip('"').strip("'") for p in parts if p.strip()]
    return _looks_like_data_row(non_empty) if non_empty else False


def _looks_like_data_row(cells: list[str]) -> bool:
    """Return True if a row already parses as a transaction record.

    A genuine data row carries both a parseable date and a parseable amount; a
    header row carries neither (``Date``/``Amount`` are labels, not values).
    Used to detect headerless files before the first row is consumed as a
    header. Reuses ``detect_date_format`` and ``parse_amount_str`` so date and
    amount recognition stay coherent with the rest of the tabular pipeline.

    Args:
        cells: Non-empty, unquoted cell strings from one row.

    Returns:
        True when at least one cell is a date and at least one is an amount.
    """
    has_date = any(detect_date_format([c])[0] is not None for c in cells)
    has_amount = any(_is_amount(c) for c in cells)
    return has_date and has_amount


def _looks_like_header(cells: list[str]) -> bool:
    """Return True if a row reads as a header rather than a data record.

    A header carries mostly non-amount labels (``Date``, ``Amount``,
    ``Description``), so few of its cells parse as amounts.

    Args:
        cells: Non-empty, unquoted cell strings from one row.

    Returns:
        True when the row has at least two cells and fewer than half parse as
        amounts.
    """
    if len(cells) < 2:
        return False
    amount_count = sum(1 for c in cells if _is_amount(c))
    return amount_count / len(cells) < 0.5


def _is_amount(s: str) -> bool:
    """Return True if the string parses as a transaction amount.

    Reuses ``parse_amount_str`` — the importer's amount parser — so header
    detection recognizes exactly the formats the loader does: parentheses
    negatives (``(42.50)``), DR/CR suffixes, currency symbols, and thousands
    separators, not a narrower float-only subset. Probes with the ``us``
    convention (as ``column_mapper`` does); the boolean result is
    format-agnostic, since european-formatted values still parse non-None.

    Args:
        s: Cell string to test.

    Returns:
        True if the cell parses as an amount.
    """
    return parse_amount_str(s, "us") is not None


def _remove_trailing_rows(
    df: pl.DataFrame, patterns: list[str]
) -> tuple[pl.DataFrame, int]:
    """Remove trailing rows matching regex patterns, scanning from end.

    Args:
        df: Input DataFrame.
        patterns: List of regex patterns to match against row values.

    Returns:
        Tuple of (trimmed DataFrame, number of rows removed).
    """
    if len(df) == 0 or not patterns:
        return df, 0

    compiled = [re.compile(p, re.IGNORECASE) for p in patterns]
    first_col = df.columns[0]
    values = df[first_col].to_list()

    remove_from = len(values)
    for i in range(len(values) - 1, -1, -1):
        val = str(values[i]) if values[i] is not None else ""
        row_str = ",".join(
            str(df[col][i]) if df[col][i] is not None else "" for col in df.columns
        )
        if any(p.search(val) or p.search(row_str) for p in compiled):
            remove_from = i
        else:
            break

    if remove_from < len(values):
        removed = len(values) - remove_from
        return df.head(remove_from), removed
    return df, 0


def _remove_repeated_headers(df: pl.DataFrame) -> pl.DataFrame:
    """Remove rows that duplicate the header row (paginated exports).

    Args:
        df: Input DataFrame.

    Returns:
        DataFrame with repeated header rows removed.
    """
    if len(df) == 0:
        return df
    headers_lower = [c.lower() for c in df.columns]
    mask = [True] * len(df)
    first_col = df.columns[0]
    first_col_values = df[first_col].cast(pl.Utf8).to_list()
    for i, val in enumerate(first_col_values):
        if val is not None and val.lower() == headers_lower[0]:
            row_values = [
                str(df[col][i]).lower() if df[col][i] is not None else ""
                for col in df.columns
            ]
            if row_values == headers_lower:
                mask[i] = False
    return df.filter(pl.Series(mask))


def _read_excel(
    path: Path,
    info: FormatInfo,
    *,
    skip_rows: int | None = None,
    sheet: str | None = None,
    source_bytes: bytes | None = None,
) -> ReadResult:
    """Read an Excel (.xlsx) file.

    Args:
        path: File path.
        info: Format detection result (unused for Excel, kept for API consistency).
        skip_rows: Rows to skip after header detection.
        sheet: Sheet name to read. If None, picks the sheet with the most rows.
        source_bytes: Already materialized workbook object to parse.

    Returns:
        ReadResult with the parsed DataFrame and sheet metadata.
    """
    import openpyxl

    sheet_used = sheet
    if sheet_used is None:
        wb = openpyxl.load_workbook(
            path if source_bytes is None else BytesIO(source_bytes),
            read_only=True,
            data_only=True,
        )
        try:
            best_sheet = wb.sheetnames[0]
            best_rows = 0
            for name in wb.sheetnames:
                ws = wb[name]
                row_count = ws.max_row or 0
                if row_count > best_rows:
                    best_rows = row_count
                    best_sheet = name
            sheet_used = best_sheet
        finally:
            wb.close()

    df = pl.read_excel(
        path if source_bytes is None else BytesIO(source_bytes),
        sheet_name=sheet_used,
        infer_schema_length=0,
    )

    actual_skip = 0
    if skip_rows is not None and skip_rows > 0:
        df = df.slice(skip_rows)
        actual_skip = skip_rows

    # pl.read_excel always consumes the sheet's first row as the header (no
    # headerless detection exists for Excel), so has_header stays True. That
    # unconditional assumption is exactly the unguarded case
    # header_row_looks_like_data protects: if the consumed header (the column
    # names) itself parses as a transaction, a real row-0 record was eaten.
    header_row_looks_like_data = _looks_like_data_row([
        str(c) for c in df.columns if str(c).strip()
    ])
    return ReadResult(
        df=df,
        skip_rows=actual_skip,
        sheet_used=sheet_used,
        header_row_looks_like_data=header_row_looks_like_data,
    )


def _read_parquet(path: Path, *, source_bytes: bytes | None = None) -> ReadResult:
    """Read a Parquet file.

    Args:
        path: File path.
        source_bytes: Already materialized Parquet object to parse.

    Returns:
        ReadResult with the parsed DataFrame.
    """
    df = pl.read_parquet(path if source_bytes is None else BytesIO(source_bytes))
    # Columnar formats carry column names in schema metadata — no data row is
    # consumed as a header, so has_header=False keeps rows_in_file == len(df).
    return ReadResult(df=df, has_header=False)


def _read_feather(path: Path, *, source_bytes: bytes | None = None) -> ReadResult:
    """Read a Feather/Arrow IPC file.

    Args:
        path: File path.
        source_bytes: Already materialized Feather object to parse.

    Returns:
        ReadResult with the parsed DataFrame.
    """
    df = pl.read_ipc(path if source_bytes is None else BytesIO(source_bytes))
    # Schema-typed like parquet — column names are metadata, not a header row.
    return ReadResult(df=df, has_header=False)
