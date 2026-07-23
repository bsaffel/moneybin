"""Validated, immutable local export publication."""

from __future__ import annotations

import csv
import fcntl
import hashlib
import json
import os
import shutil
import stat
from collections.abc import Generator
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import UTC
from pathlib import Path, PurePosixPath
from typing import Literal, cast
from uuid import uuid4
from zipfile import ZIP_DEFLATED, ZipFile

import duckdb
from openpyxl import load_workbook

from moneybin.exports.manifest import (
    LocalExportFormat,
    build_local_manifest,
    bundle_table_path,
)
from moneybin.exports.models import ExportDestination, ExportReceipt
from moneybin.exports.renderers import (
    decode_csv_cell,
    normalize_tabular_cell,
    render_csv,
    render_parquet,
    render_xlsx,
    workbook_worksheet_names,
)
from moneybin.exports.snapshot import PreparedExport, PreparedTable
from moneybin.services.request_lifetime import (
    RequestLifetime,
    current_request_lifetime,
    publication_barrier,
)

_BUNDLE_SIDECARS = {
    "manifest.json",
    "checksums.sha256",
    "data-dictionary.json",
}


@dataclass(frozen=True, slots=True)
class _ValidatedArtifact:
    checksums: dict[str, str]
    directories: tuple[Path, ...]
    files: tuple[Path, ...]


class LocalExportPublisher:
    """Render, validate, and atomically publish profile-scoped artifacts."""

    def __init__(
        self,
        profile_exports_dir: Path,
        *,
        destination_name: str = "local:exports",
    ) -> None:
        """Bind the resolved local destination without creating it."""
        self._exports_root = profile_exports_dir.expanduser().resolve()
        self._destination_name = destination_name

    def publish(
        self,
        snapshot: PreparedExport,
        *,
        format: LocalExportFormat,
        compress_zip: bool,
        publication_lifetime: RequestLifetime | None = None,
    ) -> ExportReceipt:
        """Publish a new artifact without replacing any completed run."""
        if format not in {"csv", "parquet", "xlsx"}:
            raise ValueError(f"Unsupported local export format: {format}")
        if format == "xlsx" and compress_zip:
            raise ValueError("XLSX is already compressed and rejects ZIP compression")

        lifetime = publication_lifetime or current_request_lifetime()
        if lifetime is not None:
            lifetime.raise_if_cancelled()
        self._create_exports_root()
        staging_root = self._exports_root / f".staging-{uuid4()}"
        staging_root.mkdir(mode=0o700)
        staging_root.chmod(0o700)
        try:
            artifact_path = self._render(snapshot, format, staging_root)
            if lifetime is not None:
                lifetime.raise_if_cancelled()
            if format == "xlsx":
                validated = validate_xlsx(artifact_path, snapshot)
            else:
                validated = validate_bundle(artifact_path, snapshot, format)
            _apply_restrictive_modes(validated)
            verified_checksums = dict(validated.checksums)

            zip_path = None
            if compress_zip:
                zip_path = staging_root / "artifact.zip"
                _write_zip(artifact_path, validated.files, zip_path)
                verified_checksums["archive.zip"] = validate_zip(
                    artifact_path, validated.files, zip_path
                )
                zip_path.chmod(0o600)

            with _publication_lock(self._exports_root):
                with publication_barrier(lifetime):
                    stem = self._available_stem(
                        snapshot,
                        workbook=format == "xlsx",
                        compressed=compress_zip,
                    )
                    if format == "xlsx":
                        final_path = self._exports_root / f"{stem}.xlsx"
                    else:
                        final_path = self._exports_root / stem
                    final_zip = None
                    published_primary = False
                    try:
                        artifact_path.rename(final_path)
                        published_primary = True
                        if zip_path is not None:
                            final_zip = self._exports_root / f"{stem}.zip"
                            zip_path.rename(final_zip)
                    except Exception:
                        if published_primary:
                            _remove_owned_artifact(final_path)
                        raise

            return ExportReceipt(
                subject=snapshot.subject.as_manifest(),
                format=format,
                redaction_mode=snapshot.redaction_mode,
                destination=ExportDestination(
                    destination_id=None,
                    name=self._destination_name,
                    kind="local",
                    local_path=self._exports_root,
                    spreadsheet_id=None,
                    managed_tab_prefix=None,
                ),
                artifact_path=final_path.resolve(),
                compressed_artifact_path=(
                    final_zip.resolve() if final_zip is not None else None
                ),
                sheets_identity=None,
                row_counts={table.name: len(table.rows) for table in snapshot.tables},
                output_classes={
                    table.name: {
                        column.name: column.data_class.value for column in table.columns
                    }
                    for table in snapshot.tables
                },
                checksums=verified_checksums,
                recovery_actions=(),
            )
        finally:
            if staging_root.exists():
                shutil.rmtree(staging_root)

    def _create_exports_root(self) -> None:
        if self._exports_root.exists():
            if not self._exports_root.is_dir():
                raise ValueError("local export destination is not a directory")
            return
        self._exports_root.mkdir(mode=0o700, parents=True)
        self._exports_root.chmod(0o700)

    def _render(
        self,
        snapshot: PreparedExport,
        format: LocalExportFormat,
        staging_root: Path,
    ) -> Path:
        if format == "csv":
            artifact_path = staging_root / "artifact"
            render_csv(snapshot, artifact_path)
            return artifact_path
        if format == "parquet":
            artifact_path = staging_root / "artifact"
            render_parquet(snapshot, artifact_path)
            return artifact_path
        render_xlsx(snapshot, staging_root)
        return staging_root / "export.xlsx"

    def _available_stem(
        self,
        snapshot: PreparedExport,
        *,
        workbook: bool,
        compressed: bool,
    ) -> str:
        timestamp = snapshot.created_at.astimezone(UTC)
        base = f"export-{timestamp:%Y%m%dT%H%M%SZ}"
        suffix = 1
        while True:
            stem = base if suffix == 1 else f"{base}-{suffix}"
            primary = self._exports_root / (f"{stem}.xlsx" if workbook else stem)
            archive = self._exports_root / f"{stem}.zip"
            if not primary.exists() and (not compressed or not archive.exists()):
                return stem
            suffix += 1


def validate_bundle(
    root: Path,
    snapshot: PreparedExport,
    format: Literal["csv", "parquet"],
) -> _ValidatedArtifact:
    """Independently validate emitted bundle bytes and receipt records."""
    table_paths = {
        table.name: bundle_table_path(table.name, format) for table in snapshot.tables
    }
    files, directories = _validated_bundle_layout(
        root,
        expected_files=_BUNDLE_SIDECARS | set(table_paths.values()),
    )
    manifest = _read_json_object(files["manifest.json"])
    dictionary = _read_json_object(files["data-dictionary.json"])
    if dictionary != snapshot.data_dictionary:
        raise ValueError("export data dictionary does not match prepared snapshot")

    table_files: dict[str, tuple[str, str]] = {}
    checksum_records: dict[str, str] = {}
    emitted_checksum_records = _read_checksum_records(files["checksums.sha256"])
    for table in snapshot.tables:
        relative_path = table_paths[table.name]
        artifact_path = files[relative_path]
        digest = _file_digest(artifact_path)
        table_files[table.name] = (relative_path, digest)
        checksum_records[relative_path] = digest
        if emitted_checksum_records.get(relative_path) != digest:
            raise ValueError("export table checksum validation failed")
        if _table_row_count(artifact_path, format) != len(table.rows):
            raise ValueError("export table row count validation failed")
        _validate_table_content(artifact_path, table, format)

    expected_manifest = build_local_manifest(
        snapshot,
        format=format,
        table_files=table_files,
    )
    if manifest != expected_manifest:
        raise ValueError("export manifest does not match prepared snapshot")
    if emitted_checksum_records != checksum_records:
        raise ValueError("export checksum records do not match validated bytes")
    return _ValidatedArtifact(
        checksums={name: _file_digest(path) for name, path in files.items()},
        directories=directories,
        files=tuple(files.values()),
    )


def validate_xlsx(path: Path, snapshot: PreparedExport) -> _ValidatedArtifact:
    """Read the workbook back and validate its visible data and receipts."""
    _validate_regular_file(path)
    worksheets = workbook_worksheet_names(snapshot)
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        expected_sheet_names = [
            *(worksheets[table.name] for table in snapshot.tables),
            "MoneyBin Manifest",
            "MoneyBin Data Dictionary",
        ]
        if workbook.sheetnames != expected_sheet_names:
            raise ValueError("XLSX worksheet names do not match prepared snapshot")
        if any(
            workbook[name].sheet_state != "visible" for name in expected_sheet_names
        ):
            raise ValueError("XLSX worksheets must all be visible")

        manifest = _parse_xlsx_json(workbook["MoneyBin Manifest"]["A2"].value)
        dictionary = _parse_xlsx_json(workbook["MoneyBin Data Dictionary"]["A2"].value)
        if dictionary != snapshot.data_dictionary:
            raise ValueError("XLSX data dictionary does not match prepared snapshot")
        expected_manifest = build_local_manifest(
            snapshot,
            format="xlsx",
            worksheets=worksheets,
        )
        if manifest != expected_manifest:
            raise ValueError("XLSX manifest does not match prepared snapshot")

        for prepared_table in snapshot.tables:
            sheet = workbook[worksheets[prepared_table.name]]
            header_cells = next(sheet.iter_rows(max_row=1))
            headers = tuple(cell.value for cell in header_cells)
            if headers != tuple(column.name for column in prepared_table.columns):
                raise ValueError("XLSX column validation failed")
            if sheet.max_row - 1 != len(prepared_table.rows):
                raise ValueError("XLSX row count validation failed")
            actual_rows = tuple(sheet.iter_rows(min_row=2, values_only=True))
            expected_rows = tuple(
                tuple(normalize_tabular_cell(value) for value in row)
                for row in prepared_table.rows
            )
            if actual_rows != expected_rows:
                raise ValueError("XLSX cell validation failed")
            for row_index, row in enumerate((headers, *actual_rows), start=1):
                for column_index, value in enumerate(row, start=1):
                    if isinstance(value, str) and value.startswith((
                        "=",
                        "+",
                        "-",
                        "@",
                    )):
                        cell = sheet.cell(row=row_index, column=column_index)
                        if cell.data_type != "s":
                            raise ValueError(
                                "XLSX formula-leading value is not literal text"
                            )
    finally:
        workbook.close()
    return _ValidatedArtifact(
        checksums={"export.xlsx": _file_digest(path)},
        directories=(),
        files=(path,),
    )


def validate_zip(
    bundle_root: Path,
    bundle_files: tuple[Path, ...],
    zip_path: Path,
) -> str:
    """Verify every archived byte against the already validated bundle."""
    expected = {
        path.relative_to(bundle_root).as_posix(): _file_digest(path)
        for path in bundle_files
    }
    with ZipFile(zip_path) as archive:
        if set(archive.namelist()) != set(expected):
            raise ValueError("ZIP does not contain the complete export bundle")
        for name, digest in expected.items():
            if hashlib.sha256(archive.read(name)).hexdigest() != digest:
                raise ValueError("ZIP member checksum validation failed")
    return _file_digest(zip_path)


def _write_zip(
    bundle_root: Path,
    bundle_files: tuple[Path, ...],
    zip_path: Path,
) -> None:
    with ZipFile(zip_path, "w", compression=ZIP_DEFLATED) as archive:
        for path in sorted(bundle_files):
            archive.write(path, path.relative_to(bundle_root).as_posix())


def _remove_owned_artifact(path: Path) -> None:
    if path.is_dir():
        shutil.rmtree(path)
    else:
        path.unlink()


@contextmanager
def _publication_lock(exports_root: Path) -> Generator[None]:
    lock_path = exports_root / ".publish.lock"
    with lock_path.open("a+b") as handle:
        lock_path.chmod(0o600)
        fcntl.flock(handle.fileno(), fcntl.LOCK_EX)
        try:
            yield
        finally:
            fcntl.flock(handle.fileno(), fcntl.LOCK_UN)


def _table_row_count(path: Path, format: Literal["csv", "parquet"]) -> int:
    if format == "csv":
        with path.open(encoding="utf-8", newline="") as handle:
            reader = csv.reader(handle)
            next(reader)
            return sum(1 for _row in reader)
    row = duckdb.read_parquet(str(path)).aggregate("count(*)").fetchone()
    if row is None:
        raise ValueError("Parquet row count validation returned no result")
    return cast(int, row[0])


def _validate_table_content(
    path: Path,
    table: PreparedTable,
    format: Literal["csv", "parquet"],
) -> None:
    """Round-trip each emitted table against its prepared schema and values."""
    expected_headers = tuple(column.name for column in table.columns)
    if format == "csv":
        with path.open(encoding="utf-8", newline="") as handle:
            rows = tuple(csv.reader(handle))
        headers, *data_rows = rows
        expected_rows = tuple(
            tuple(
                None if value is None else str(normalize_tabular_cell(value))
                for value in row
            )
            for row in table.rows
        )
        actual_rows = tuple(
            tuple(decode_csv_cell(value) for value in row) for row in data_rows
        )
    else:
        relation = duckdb.read_parquet(str(path))
        headers = tuple(relation.columns)
        actual_rows = tuple(relation.fetchall())
        expected_rows = table.rows
    actual_headers = (
        tuple(decode_csv_cell(header) for header in headers)
        if format == "csv"
        else tuple(headers)
    )
    if actual_headers != expected_headers:
        raise ValueError("export table column validation failed")
    if actual_rows != expected_rows:
        raise ValueError("export table cell validation failed")


def _validated_bundle_layout(
    root: Path,
    *,
    expected_files: set[str],
) -> tuple[dict[str, Path], tuple[Path, ...]]:
    _validate_directory(root)
    expected_directories = _expected_bundle_directories(expected_files)
    files: dict[str, Path] = {}
    directories: dict[str, Path] = {}
    pending = [root]
    while pending:
        directory = pending.pop()
        with os.scandir(directory) as entries:
            for entry in entries:
                path = Path(entry.path)
                relative_path = path.relative_to(root).as_posix()
                if entry.is_symlink():
                    raise ValueError("export bundle contains a symlink")
                if entry.is_dir(follow_symlinks=False):
                    directories[relative_path] = path
                    pending.append(path)
                elif entry.is_file(follow_symlinks=False):
                    files[relative_path] = path
                else:
                    raise ValueError("export bundle contains an unsupported entry")

    if set(files) != expected_files or set(directories) != expected_directories:
        raise ValueError("export bundle layout does not match prepared snapshot")
    ordered_directories = (root, *(directories[name] for name in sorted(directories)))
    return ({name: files[name] for name in sorted(files)}, ordered_directories)


def _expected_bundle_directories(expected_files: set[str]) -> set[str]:
    directories = {"tables"}
    for name in expected_files:
        relative = PurePosixPath(name)
        if relative.is_absolute() or ".." in relative.parts:
            raise ValueError("export bundle contains an unsafe expected path")
        parent = relative.parent
        while parent != PurePosixPath("."):
            directories.add(parent.as_posix())
            parent = parent.parent
    return directories


def _validate_directory(path: Path) -> None:
    mode = path.lstat().st_mode
    if stat.S_ISLNK(mode):
        raise ValueError("export bundle contains a symlink")
    if not stat.S_ISDIR(mode):
        raise ValueError("export bundle root is not a directory")


def _validate_regular_file(path: Path) -> None:
    mode = path.lstat().st_mode
    if stat.S_ISLNK(mode):
        raise ValueError("export artifact is a symlink")
    if not stat.S_ISREG(mode):
        raise ValueError("export artifact is not a regular file")


def _read_checksum_records(path: Path) -> dict[str, str]:
    records: dict[str, str] = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        parts = line.split("  ", maxsplit=1)
        if len(parts) != 2 or parts[1] in records:
            raise ValueError("export checksum file is malformed")
        digest, relative_path = parts
        if len(digest) != 64:
            raise ValueError("export checksum file is malformed")
        records[relative_path] = digest
    return records


def _read_json_object(path: Path) -> dict[str, object]:
    value = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(value, dict):
        raise ValueError("export JSON receipt must be an object")
    return cast(dict[str, object], value)


def _parse_xlsx_json(value: object) -> dict[str, object]:
    if not isinstance(value, str):
        raise ValueError("XLSX JSON receipt cell must contain text")
    parsed = json.loads(value)
    if not isinstance(parsed, dict):
        raise ValueError("XLSX JSON receipt must be an object")
    return cast(dict[str, object], parsed)


def _file_digest(path: Path) -> str:
    with path.open("rb") as handle:
        return hashlib.file_digest(handle, "sha256").hexdigest()


def _apply_restrictive_modes(artifact: _ValidatedArtifact) -> None:
    for directory in artifact.directories:
        directory.chmod(0o700)
    for file in artifact.files:
        file.chmod(0o600)
