"""Pure local renderers for immutable prepared exports."""

from __future__ import annotations

import csv
import hashlib
import json
import re
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal
from enum import Enum
from pathlib import Path
from typing import cast
from uuid import UUID

import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from moneybin.exports.manifest import (
    CSV_ENCODING,
    LocalExportFormat,
    build_local_manifest,
    bundle_table_path,
    safe_table_identifier,
)
from moneybin.exports.snapshot import PreparedExport, PreparedTable

_MANIFEST_SHEET = "MoneyBin Manifest"
_DICTIONARY_SHEET = "MoneyBin Data Dictionary"
_FORMULA_PREFIXES = ("=", "+", "-", "@")
_DECIMAL_TYPE = re.compile(r"^DECIMAL\((\d+),\s*(\d+)\)$")
_XLSX_EMPTY = r"\E"
_XLSX_ESCAPE = "\\"
_MAX_XLSX_CELL_TEXT_LENGTH = 16_000

type TabularCell = bool | int | float | str | None


@dataclass(frozen=True, slots=True)
class RenderedArtifact:
    """Paths and receipts produced by one pure renderer."""

    path: Path
    manifest: Mapping[str, object]
    table_files: Mapping[str, Path]
    file_checksums: Mapping[str, str]


def render_csv(snapshot: PreparedExport, staging_root: Path) -> RenderedArtifact:
    """Render one prepared snapshot as a CSV directory bundle."""
    return _render_bundle(snapshot, staging_root, format="csv")


def render_parquet(snapshot: PreparedExport, staging_root: Path) -> RenderedArtifact:
    """Render one prepared snapshot as a native Parquet directory bundle."""
    return _render_bundle(snapshot, staging_root, format="parquet")


def render_xlsx(snapshot: PreparedExport, staging_root: Path) -> RenderedArtifact:
    """Render one prepared snapshot as a workbook with visible receipts."""
    staging_root.mkdir(mode=0o700, parents=True, exist_ok=True)
    workbook = Workbook()
    active_sheet = workbook.active
    if active_sheet is None:
        raise RuntimeError("new XLSX workbook has no active sheet")
    workbook.remove(active_sheet)
    worksheets = workbook_worksheet_names(snapshot)

    for table in snapshot.tables:
        sheet_name = worksheets[table.name]
        sheet = workbook.create_sheet(sheet_name)
        _write_xlsx_row(sheet, 1, [column.name for column in table.columns])
        for row_index, row in enumerate(table.rows, start=2):
            _write_xlsx_row(
                sheet,
                row_index,
                [normalize_tabular_cell(value) for value in row],
                encode_empty_strings=True,
            )

    manifest = build_local_manifest(
        snapshot,
        format="xlsx",
        worksheets=worksheets,
    )
    manifest_sheet = workbook.create_sheet(_MANIFEST_SHEET)
    _write_xlsx_json_sheet(manifest_sheet, manifest)
    dictionary_sheet = workbook.create_sheet(_DICTIONARY_SHEET)
    _write_xlsx_json_sheet(dictionary_sheet, snapshot.data_dictionary)

    workbook_path = staging_root / "export.xlsx"
    workbook.save(workbook_path)
    workbook_path.chmod(0o600)
    return RenderedArtifact(
        path=workbook_path,
        manifest=manifest,
        table_files={},
        file_checksums={"export.xlsx": _file_digest(workbook_path)},
    )


def _render_bundle(
    snapshot: PreparedExport,
    staging_root: Path,
    *,
    format: LocalExportFormat,
) -> RenderedArtifact:
    if format == "xlsx":
        raise ValueError("XLSX is a workbook, not a directory bundle")
    staging_root.mkdir(mode=0o700, parents=True, exist_ok=True)
    staging_root.chmod(0o700)
    tables_root = staging_root / "tables"
    tables_root.mkdir(mode=0o700)
    tables_root.chmod(0o700)
    table_files: dict[str, Path] = {}
    relative_files: dict[str, tuple[str, str]] = {}

    for table in snapshot.tables:
        relative_path = bundle_table_path(table.name, format)
        table_path = staging_root.joinpath(*Path(relative_path).parts)
        if table_path.exists():
            raise ValueError("prepared table names produce duplicate artifact paths")
        if format == "csv":
            _write_csv(table, table_path)
        else:
            _write_parquet(table, table_path)
        table_path.chmod(0o600)
        digest = _file_digest(table_path)
        table_files[table.name] = table_path
        relative_files[table.name] = (relative_path, digest)

    manifest = build_local_manifest(
        snapshot,
        format=format,
        table_files=relative_files,
    )
    _write_json(staging_root / "manifest.json", manifest)
    _write_json(staging_root / "data-dictionary.json", snapshot.data_dictionary)
    checksum_lines = [
        f"{digest}  {relative_path}\n"
        for relative_path, digest in sorted(relative_files.values())
    ]
    checksums_path = staging_root / "checksums.sha256"
    checksums_path.write_text("".join(checksum_lines), encoding="utf-8", newline="")
    checksums_path.chmod(0o600)
    return RenderedArtifact(
        path=staging_root,
        manifest=manifest,
        table_files=table_files,
        file_checksums=dict(relative_files.values()),
    )


def _write_csv(table: PreparedTable, path: Path) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, lineterminator="\n")
        writer.writerow(_csv_payload_cell(column.name) for column in table.columns)
        writer.writerows(
            [_csv_payload_cell(value) for value in row] for row in table.rows
        )


def _write_parquet(table: PreparedTable, path: Path) -> None:
    schema = parquet_schema_for(table)
    arrow_table = pa.Table.from_arrays(
        [
            pa.array([row[index] for row in table.rows], type=field.type)
            for index, field in enumerate(schema)
        ],
        schema=schema,
    )
    pq.write_table(arrow_table, path)  # type: ignore[reportUnknownMemberType]  # pyarrow lacks complete type stubs


def parquet_schema_for(table: PreparedTable) -> pa.Schema:
    """Build the exact Arrow schema promised by prepared DuckDB columns."""
    return pa.schema([
        pa.field(column.name, _arrow_type(column.duckdb_type))
        for column in table.columns
    ])


def _arrow_type(duckdb_type: str) -> pa.DataType:
    """Map the closed prepared-export DuckDB type vocabulary to Arrow."""
    source = duckdb_type.strip()
    normalized = source.upper()
    if source.endswith("[]"):
        return pa.list_(_arrow_type(source[:-2]))
    if normalized.startswith("LIST(") and source.endswith(")"):
        return pa.list_(_arrow_type(source[5:-1]))
    if normalized.startswith("STRUCT(") and source.endswith(")"):
        return pa.struct([
            _arrow_struct_field(field)
            for field in _split_duckdb_type_fields(source[7:-1])
        ])
    simple_types = {
        "BOOLEAN": pa.bool_(),
        "TINYINT": pa.int8(),
        "SMALLINT": pa.int16(),
        "INTEGER": pa.int32(),
        "BIGINT": pa.int64(),
        "UTINYINT": pa.uint8(),
        "USMALLINT": pa.uint16(),
        "UINTEGER": pa.uint32(),
        "UBIGINT": pa.uint64(),
        "FLOAT": pa.float32(),
        "DOUBLE": pa.float64(),
        "VARCHAR": pa.string(),
        "TEXT": pa.string(),
        "STRING": pa.string(),
        "DATE": pa.date32(),
        "TIMESTAMP": pa.timestamp("us"),
        "TIMESTAMP WITH TIME ZONE": pa.timestamp("us", tz="UTC"),
    }
    if normalized in simple_types:
        return simple_types[normalized]
    decimal = _DECIMAL_TYPE.fullmatch(normalized)
    if decimal is not None:
        return pa.decimal128(int(decimal.group(1)), int(decimal.group(2)))
    raise ValueError(f"Unsupported Parquet column type: {duckdb_type}")


def _arrow_struct_field(field: str) -> pa.Field:
    """Map one DuckDB ``STRUCT(name TYPE)`` field to Arrow."""
    name, separator, duckdb_type = field.strip().partition(" ")
    if not separator or not duckdb_type.strip():
        raise ValueError(f"Unsupported Parquet struct field: {field}")
    return pa.field(name.strip('"'), _arrow_type(duckdb_type))


def _split_duckdb_type_fields(value: str) -> list[str]:
    """Split comma-separated nested DuckDB type fields at their outer level."""
    fields: list[str] = []
    start = 0
    depth = 0
    for index, character in enumerate(value):
        if character == "(":
            depth += 1
        elif character == ")":
            depth -= 1
        elif character == "," and depth == 0:
            fields.append(value[start:index])
            start = index + 1
    fields.append(value[start:])
    return fields


def _csv_payload_cell(value: object) -> object:
    if value is None:
        return CSV_ENCODING["null"]
    if isinstance(value, Mapping):
        return _json_text(_payload_json_safe(cast(Mapping[object, object], value)))
    if isinstance(value, (list, tuple)):
        return _json_text(_payload_json_safe(cast(Sequence[object], value)))
    normalized = normalize_tabular_cell(value)
    if isinstance(value, str) and value.startswith(
        tuple(cast(list[str], CSV_ENCODING["escaped_prefixes"]))
    ):
        return f"{CSV_ENCODING['escape']}{normalized}"
    return normalized


def decode_csv_cell(value: str | None) -> str | None:
    """Decode one versioned CSV text cell after the reader applies the null marker."""
    if value is None or value == CSV_ENCODING["null"]:
        return None
    escape = cast(str, CSV_ENCODING["escape"])
    escaped_prefixes = tuple(cast(list[str], CSV_ENCODING["escaped_prefixes"]))
    if value.startswith(escape):
        decoded = value[len(escape) :]
        if not decoded.startswith(escaped_prefixes):
            raise ValueError("CSV cell contains an invalid escape sequence")
        return decoded
    return value


def normalize_tabular_cell(value: object) -> TabularCell:
    """Normalize a prepared cell for lossless text-oriented payloads."""
    if value is None or isinstance(value, (bool, int, float, str)):
        return value
    if isinstance(value, Decimal):
        # Text preserves DECIMAL precision; XLSX numeric cells are IEEE-754 doubles.
        return str(value)
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, bytes):
        return value.hex()
    if isinstance(value, (UUID, Path, Enum)):
        return str(value)
    if isinstance(value, Mapping):
        return _json_text(_payload_json_safe(cast(Mapping[object, object], value)))
    if isinstance(value, (list, tuple)):
        return _json_text(_payload_json_safe(cast(Sequence[object], value)))
    raise TypeError(f"Unsupported export cell value: {type(value).__name__}")


def _payload_json_safe(value: object) -> object:
    if isinstance(value, Mapping):
        mapping = cast(Mapping[object, object], value)
        return {str(key): _payload_json_safe(item) for key, item in mapping.items()}
    if isinstance(value, (list, tuple)):
        sequence = cast(Sequence[object], value)
        return [_payload_json_safe(item) for item in sequence]
    return normalize_tabular_cell(value)


def _json_text(value: object) -> str:
    return json.dumps(
        value,
        ensure_ascii=False,
        separators=(",", ":"),
        sort_keys=True,
    )


def _write_json(path: Path, value: object) -> None:
    path.write_text(f"{_json_text(value)}\n", encoding="utf-8", newline="")
    path.chmod(0o600)


def _file_digest(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def workbook_worksheet_names(snapshot: PreparedExport) -> dict[str, str]:
    """Return deterministic Excel-safe names with case-insensitive uniqueness."""
    used = {_MANIFEST_SHEET.casefold(), _DICTIONARY_SHEET.casefold()}
    worksheets: dict[str, str] = {}
    for table in snapshot.tables:
        sheet_name = _worksheet_name(table.name, used)
        used.add(sheet_name.casefold())
        worksheets[table.name] = sheet_name
    return worksheets


def _worksheet_name(name: str, used_casefolded: set[str]) -> str:
    base = safe_table_identifier(name)
    base = base[:31]
    candidate = base
    suffix = 2
    while candidate.casefold() in used_casefolded:
        marker = f"-{suffix}"
        candidate = f"{base[: 31 - len(marker)]}{marker}"
        suffix += 1
    return candidate


def _write_xlsx_row(
    sheet: Worksheet,
    row: int,
    values: Sequence[TabularCell],
    *,
    encode_empty_strings: bool = False,
) -> None:
    for column, raw_value in enumerate(values, start=1):
        value = _xlsx_payload_cell(raw_value) if encode_empty_strings else raw_value
        cell = sheet.cell(row=row, column=column, value=value)
        if isinstance(value, str) and value.startswith(_FORMULA_PREFIXES):
            cell.data_type = "s"


def _write_xlsx_json_sheet(sheet: Worksheet, value: object) -> None:
    """Write receipt JSON in cells that stay below Excel's text limit."""
    sheet.append(["JSON"])
    text = _json_text(value)
    for start in range(0, len(text), _MAX_XLSX_CELL_TEXT_LENGTH):
        sheet.append([text[start : start + _MAX_XLSX_CELL_TEXT_LENGTH]])


def _xlsx_payload_cell(value: TabularCell) -> TabularCell:
    """Encode empty strings without conflating them with blank XLSX cells."""
    if value == "":
        return _XLSX_EMPTY
    if isinstance(value, str) and value.startswith(_XLSX_ESCAPE):
        return f"{_XLSX_ESCAPE}{value}"
    return value


def decode_xlsx_cell(value: object) -> object:
    """Decode the reversible XLSX empty-string representation."""
    if value == _XLSX_EMPTY:
        return ""
    if isinstance(value, str) and value.startswith(f"{_XLSX_ESCAPE}{_XLSX_ESCAPE}"):
        return value[1:]
    return value
